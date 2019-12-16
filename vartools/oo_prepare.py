"""
This module prepares oocyte data formatted in the CFERV 'daily recording files' format

main function:
    convert_all(filelist, header, outddir)

"""

from openpyxl import load_workbook
import datetime
from string import ascii_uppercase
import re
from os.path import splitext as path_splitext
from sys import exit as sys_exit

# my oocyte fitting project
from pydrc import fit4pdrc as fit

def readrow(ws, cols, i):
    row = []
    for letter in cols:
        val = ws[letter + str(i)].value
        row.append(val)
    return(row)

def xl2list(xlfile):
    """
    creates a list of values for each row in the oocyte data file
    """
    # starting values
    start = 3
    oononelist = [None] * 26
    oorow = ''
    expnonelist = [None] * 10
    exprow = ''
    #prefix = [""]
    # end starting values
    wb = load_workbook(xlfile)
    ws = wb['DataEntry']
    assay = ws['AA3'].value
    if not assay:
        sys_exit("\nAssay code was not defined, exiting now")
    if assay not in ("gluDRC","glyDRC","mgDRC","znDRC","pH","pHDRC","MTSEA"):
        lassay = assay.lower().rstrip()
        if lassay == "gludrc":
            print("--- Correcting assay code from " + assay, end='')
            assay = "gluDRC"
            ws['AA3'].value = assay
            print(" to " + assay)
            print(xlfile, end = '')
        elif lassay == "glydrc":
            print("--- Correcting assay code from " + assay, end='')
            assay = "glyDRC"
            ws['AA3'].value = assay
            print(" to " + assay)
            print(xlfile, end = '')
        elif lassay == "mgdrc":
            print("--- Correcting assay code from " + assay, end='')
            assay = "mgDRC"
            ws['AA3'].value = assay
            print(" to " + assay)
            print(xlfile, end = '')
        elif lassay == "zndrc":
            print("--- Correcting assay code from " + assay, end='')
            assay = "znDRC"
            ws['AA3'].value = assay
            print(" to " + assay)
            print(xlfile, end = '')
        elif lassay == "ph":
            print("--- Correcting assay code from " + assay, end='')
            assay = "pH"
            ws['AA3'].value = assay
            print(" to " + assay)
            print(xlfile, end = '')
        elif lassay == "mtsea":
            print("--- Correcting assay code from " + assay, end='')
            assay = "MTSEA"
            ws['AA3'].value = assay
            print(" to " + assay)
            print(xlfile, end = '')
        elif lassay == "phdrc":
            print("--- Correcting assay code from " + assay, end='')
            assay = "pHDRC"
            ws['AA3'].value = assay
            print(" to " + assay)
            print(xlfile, end = '')
        else:
            print("\nWARNING: Assay code '" + assay + "' is not currently supported by the database")
            print("Are you sure you know what you're doing?")
            user_input = input("Continue? (yes/no/ENTER=no)")
            if not user_input:
                sys_exit()
            elif user_input == "no":
                sys_exit()
            elif user_intput == "yes":
                pass
            else:
                sys_exit()
    i = start
    # the column names of the oocyte data happends to be uppercase letters A-Z
    oocols = ascii_uppercase
    oolist = []
    # continue to read data until the last row is blank
    fitlist = []
    while oorow != oononelist:
        oorow = readrow(ws, oocols, i)
        if oorow[0] == "display mode":
            # delete any display mode data
            i += 1
            continue
        if oorow[-1] == "DEL":
            # delete this data, it will not make it into the database
            i += 1
            continue
        if oorow[-1] == "ASSAY=1min1conc":
            # delete this data for now
            i += 1
            continue
        if oorow[-1] == "NOFIT":
            # this code is used for data that cannot be fit
            fitlist.append([None,None,None,None,None,None])
            oolist.append(oorow)
            i += 1
            continue
        # lets fit the data
        data = oorow[5:-2]
        #print(data)
        doses = [-10,-9.52,-9,-8.52,-8,-7.52,-7,-6.52,-6,-5.52,
                 -5,-4.52,-4,-3.52,-3,-2.52,-2,-1.52,-1]
        fit_data = []
        fit_dose = []
        for n, value in enumerate(data):
            if value != None:
                fit_data.append(value)
                fit_dose.append(doses[n])
        if assay in ('mgDRC','znDRC'):
            top_var = True
        else:
            top_var = False
        ret = fit(fit_dose, fit_data, top_const=top_var)
        #print(ret)
        fits = [ret['c'],ret['h'],ret['b'],ret['t'],ret['p'],ret['i']]
        # below adds the fits back into the workbook
        ws['AO'+str(i)].value = round(ret['b'],2)
        ws['AP'+str(i)].value = round(ret['t'],2)
        ws['AQ'+str(i)].value = round(ret['c'],2)
        ws['AR'+str(i)].value = round(ret['h'],2)
        ws['AS'+str(i)].value = 10**round(ret['c'],2)
        fitlist.append(fits)
        oolist.append(oorow)
        i += 1
    oolist.pop(-1) # the last row will be blank, need to remove it
    i = start #3
    expcols = ['AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']
    explist = []
    while exprow != expnonelist:
        exprow = readrow(ws, expcols, i)
        #print(exprow)
        date_rec = exprow[2]
        if (exprow == expnonelist):
            continue
        date_inj = exprow[1]
        if (not (isinstance(date_rec, datetime.datetime))):
            sys_exit('\nInvalid type/value for date_rec' + str(date_rec))
        if (not (isinstance(date_inj, datetime.datetime))):
            sys_exit('\nInvlaid type/value for date_inj' + str(date_inj))
        if (not (date_rec <= datetime.datetime.now())):
            sys_exit('\nAre these recordings from the future? Your date_rec is: ' + date_rec.strftime("%Y%m%d"))
        if (not (date_inj <= datetime.datetime.now())):
            sys_exit('\nAre these injections from the future? Your date_inj is: ' + date_inj.strftime("%Y%m%d"))
        if (not (date_rec > date_inj)):
            sys_exit('\nYour recorded your oocytes before you injected them? How did you manage that? (your date_rec happened before your date_inj)')
        if (not (datetime.date(year=2012, month=1, day=1) < date_rec.date())):
            sys_exit('\nI doubt you did these recordings on ' + date_rec.strftime("%Y%m%d"))
        if (not (datetime.date(year=2012, month=1, day=1) < date_inj.date())):
            sys_exit('\nI doubt you did these recordings on ' + date_inj.strftime("%Y%m%d"))
        date_delta = date_rec - date_inj
        if (date_delta > datetime.timedelta(days=15)):
            sys_exit('\nI doubt that you recorded more than 15 days after injections...')
        explist.append(exprow)
        i += 1
    #explist.pop(-1)
    dlist = []
    if len(explist) == 1:
        for i,row in enumerate(oolist):
            newrow = row + fitlist[i] + explist[0]
            dlist.append(newrow)
    elif len(explist) == len(oolist):
        for i in range(0, len(oolist)):
            newrow = oolist[i] + fitlist[i] + explist[i]
            dlist.append(newrow)
    else:
        explist = explist[0]
        for i,row in enumerate(oolist):
            newrow = row + fitlist[i] + explist
            dlist.append(newrow)
    wb.save(xlfile)
    return(dlist)

def writeoofile(dlist, header, xlfile, outdir):
    """
    writes a .csv file in the specified format
    """
    with open(header, 'r') as hfile:
        header = hfile.readlines()
    hfile.close()
    oofile = dlist[0][0]
    p = re.compile(".+-.+-.+-.+-")
    res = p.findall(oofile)
    ph = re.compile("p[hH]")
    phres = ph.search(oofile)
    if (phres):
        oofile = outdir + '/' + res[0][:-1] + '.csv'
    else:
        oofile = outdir + '/' + res[0][:-1] + '.csv'
    with open(oofile, 'w') as opfile:
        for row in header:
            opfile.write(row)
        newdlist = []
        for row in dlist:
            newrow = []
            for value in row:
                if type(value) is int:
                    newval = str(value)
                elif type(value) is float:
                    newval = str(value)
                elif value is None:
                    newval = ""
                elif isinstance(value, datetime.datetime):
                    newval = value.strftime("%Y-%m-%d")
                else:
                    newval = value.replace(",",";")
                newrow.append(newval)
            newdlist.append(newrow)
            opfile.write(",".join(newrow) + "\n")
    opfile.close()
    return(None)

def ooconv(xlfile, header, outdir):
    """
    converts a single excel file into .csv files in the specified outdir
    """
    dlist = xl2list(xlfile)
    writeoofile(dlist, header, xlfile, outdir)
    print(" --- csv file sucessfully written to disk")
    return(None)

def convert_all(filelist, header, outdir):
    """ converts a list of excel files into .csv files in the specified outdir """
    for xlfile in filelist:
        print(xlfile, end='')
        pre, ext = path_splitext(xlfile)
        if ext not in (".xlsx",".xls"):
            print("invlaid filetype: " + ext)
            continue
        ooconv(xlfile, header, outdir)
