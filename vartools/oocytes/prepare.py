"""
This module prepares oocyte data formatted in the CFERV 'daily recording files' format

main function:
    convert_all(filelist, header, outddir)

"""

from openpyxl import load_workbook
import datetime
from string import ascii_uppercase
import re
from os.path import splitext as path_splitext
from sys import exit as sys_exit

# my oocyte fitting project
from pydrc import fit4pdrc as fit

def readrow(ws, cols, i):
    row = []
    for letter in cols:
        val = ws[letter + str(i)].value
        row.append(val)
    return(row)

def xl2list(xlfile):
    """
    creates a list of values for each row in the oocyte data file
    """
    # starting values
    start = 3
    oononelist = [None] * 26
    oorow = ''
    expnonelist = [None] * 10
    exprow = ''
    #prefix = [""]
    # end starting values
    wb = load_workbook(xlfile)
    ws = wb['DataEntry']
    i = start
    # the column names of the oocyte data happends to be uppercase letters A-Z
    oocols = ascii_uppercase
    oolist = []
    # continue to read data until the last row is blank
    fitlist = []
    while oorow != oononelist:
        oorow = readrow(ws, oocols, i)
        if oorow[-1] == "DEL":
            i += 1
            continue
        # lets fit the data
        data = oorow[5:-2]
        #print(data)
        doses = [-10,-9.52,-9,-8.52,-8,-7.52,-7,-6.52,-6,-5.52,
                 -5,-4.52,-4,-3.52,-3,-2.52,-2,-1.52,-1]
        fit_data = []
        fit_dose = []
        for n, value in enumerate(data):
            if value != None:
                fit_data.append(value)
                fit_dose.append(doses[n])
        ret = fit(fit_dose, fit_data)
        #print(ret)
        fits = [ret['c'],ret['h'],ret['b'],ret['t'],ret['p'],ret['i']]
        fitlist.append(fits)
        oolist.append(oorow)
        i += 1
    oolist.pop(-1) # the last row will be blank, need to remove it
    i = start
    expcols = ['AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ']
    explist = []
    while exprow != expnonelist:
        exprow = readrow(ws, expcols, i)
        explist.append(exprow)
        i += 1
    explist.pop(-1)
    dlist = []
    if len(explist) == 1:
        for i,row in enumerate(oolist):
            newrow = row + fitlist[i] + explist[0]
            dlist.append(newrow)
    elif len(explist) == len(oolist):
        for i in range(0, len(oolist)):
            newrow = oolist[i] + fitlist[i] + explist[i]
            dlist.append(newrow)
    else:
        explist = explist[0]
        for i,row in enumerate(oolist):
            newrow = row + fitlist[i] + explist
            dlist.append(newrow)
    return(dlist)

def writeoofile(dlist, header, xlfile, outdir):
    """
    writes a .csv file in the specified format
    """
    with open(header, 'r') as hfile:
        header = hfile.readlines()
    hfile.close()
    oofile = dlist[0][0]
    p = re.compile(".+-.+-.+-.+-")
    res = p.findall(oofile)
    ph = re.compile("p[hH]")
    phres = ph.search(oofile)
    if (phres):
        oofile = outdir + '/' + res[0][:-1] + '.csv'
    else:
        oofile = outdir + '/' + res[0][:-1] + '.csv'
    with open(oofile, 'w') as opfile:
        for row in header:
            opfile.write(row)
        newdlist = []
        for row in dlist:
            newrow = []
            for value in row:
                if type(value) is int:
                    newval = str(value)
                elif type(value) is float:
                    newval = str(value)
                elif value is None:
                    newval = ""
                elif isinstance(value, datetime.datetime):
                    newval = value.strftime("%Y-%m-%d")
                else:
                    newval = value.replace(",",";")
                newrow.append(newval)
            newdlist.append(newrow)
            opfile.write(",".join(newrow) + "\n")
    opfile.close()
    return(None)

def ooconv(xlfile, header, outdir):
    """
    converts a single excel file into .csv files in the specified outdir
    """
    dlist = xl2list(xlfile)
    writeoofile(dlist, header, xlfile, outdir)
    print("csv file sucessfully written to disk")
    return(None)

def convert_all(filelist, header, outdir):
    """ converts a list of excel files into .csv files in the specified outdir """
    for xlfile in filelist:
        print(xlfile)
        pre, ext = path_splitext(xlfile)
        if ext not in (".xlsx",".xls"):
            print("invlaid filetype: " + ext)
            continue
        ooconv(xlfile, header, outdir)
